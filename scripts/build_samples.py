"""Generate the sample RFQ files used by the tests and by the UI upload path.

Three files, each awkward in a different way:

  clean-bom.xlsx   the easy case -- header in row 1, obvious column names
  messy-bom.xlsx   header four rows down under a title block, merged cells,
                   quantity to the LEFT of the part number, a decoy sheet with
                   more rows than the real one, and price columns that look
                   like counts
  rfq-email.txt    an email body with a signature, a quoted reply chain, a
                   numbered list and quantities written four different ways

Run: python scripts/build_samples.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "sample-rfqs"

# Real MPNs from the seeded catalog, so the samples resolve end to end.
CLEAN_ROWS = [
    ("STM32F130C3Y6", "STMicroelectronics", 500, 3.98),
    ("SN74HCT08DBVR", "Texas Instruments", 2500, 0.84),
    ("CRCW0603100KJKEA", "Vishay", 10000, 0.0146),
    ("MMSZ4744AT1G", "onsemi", 2000, 0.0176),
    ("LD1117V18", "STMicroelectronics", 750, 0.42),
    ("PSMN10V4-60YL", "Nexperia", 250, 7.87),
]

MESSY_ROWS = [
    (500, "296-STM32F130C3Y6-ND", "MCU 32-bit", 3.98),
    (250, "psmn10v4-60yl-tr", "Power MOSFET", 7.87),
    (1000, "SST25VF032B-75I", "Serial flash", 1.39),
    (75, "MCP1827-0180E", "LDO regulator", 1.79),
    (5000, "BZX84-C10,215", "Zener diode", 0.0161),
    (100, "adp6208armz-2.5", "LDO regulator", 7.66),
    (12, "D38999/24JA103SN", "MIL circular connector", 7.79),
]

EMAIL = """From: dana.whitfield@acme-aero.example
To: quotes@brokerhouse.example
Subject: RFQ 41822 - urgent, line down
Date: Tue, 11 Aug 2026 09:14:02 +0100

Hi,

We have a line down and need the following as soon as you can. Target
pricing where shown, but availability matters more than price right now.

1. 296-STM32F130C3Y6-ND x 500
2. stm32f105kct7, 250, target $3.10
3. MMSZ4744ATIG QTY: 2000
4. D38999/24JA103SN\t12

- CRCW 0603 100K JKEA  10,000 pcs
- SST25VF032B-75I  1,000 pcs
- AD2736BRZ-REEL x 40
- BZX84-C10,215 x 5000

We are also told these three are getting tight, please advise:
PSMN10V4-60YL-TR x 250
adp6208armz-2.5, 100
MCP6001-E/SN  500 pcs

And one we cannot identify from the old drawing: ZZQQ99NOTAPART x 5

Let me know on lead times and whether any need full AS6171 testing.

Thanks,
Dana Whitfield
Senior Buyer, Acme Aerospace
+44 20 7946 0958
dana.whitfield@acme-aero.example

> On Tue, 11 Aug 2026, procurement wrote:
> Please also check whether the obsolete lines can be sourced authorised.
> We previously bought LM317T x 9999 but that is no longer relevant.
"""


def build_clean_bom(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    headers = ["Part Number", "Manufacturer", "Qty", "Target Price"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in CLEAN_ROWS:
        sheet.append(list(row))
    workbook.save(path)


def build_messy_bom(path: Path) -> None:
    workbook = Workbook()

    # Decoy sheet first, and with MORE rows than the real one, so a parser that
    # picks the first or the largest sheet gets it wrong.
    notes = workbook.active
    notes.title = "Instructions"
    for line in [
        ["ACME AEROSPACE - SUPPLIER INSTRUCTIONS"],
        ["Revision 4.2, issued 2026-02-01"],
        [""],
        ["1. All parts must be supplied with full traceability to the OCM."],
        ["2. Quote lead time in weeks from PO receipt."],
        ["3. Reference standard AS6171 for counterfeit test methods."],
        ["4. Payment terms Net 45 unless agreed otherwise."],
        ["5. Contact procurement@acme-aero.example with questions."],
        [""],
        ["Do not quote from this sheet. The bill of materials is on tab 'Rev C'."],
    ]:
        notes.append(line)

    sheet = workbook.create_sheet("Rev C")
    # Title block: the table does not start at row 1.
    sheet["A1"] = "ACME AEROSPACE"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells("A1:D1")
    sheet["A2"] = "Project Kestrel - Bill of Materials"
    sheet.merge_cells("A2:D2")
    sheet["A3"] = "Rev C / 2026-08-11"

    # Header on row 5, quantity to the LEFT of the part number.
    sheet["A5"] = "Required"
    sheet["B5"] = "Component"
    sheet["C5"] = "Description"
    sheet["D5"] = "Last Paid (USD)"
    for cell in sheet[5]:
        if cell.value:
            cell.font = Font(bold=True)

    for index, row in enumerate(MESSY_ROWS, start=6):
        sheet[f"A{index}"] = row[0]
        sheet[f"B{index}"] = row[1]
        sheet[f"C{index}"] = row[2]
        sheet[f"D{index}"] = row[3]

    # A merged note spanning the table, mid-data -- fills several cells that
    # would otherwise be blank.
    last = 6 + len(MESSY_ROWS)
    sheet[f"A{last + 1}"] = "All lines NCNR. Obsolete parts require test report."
    sheet.merge_cells(f"A{last + 1}:D{last + 1}")

    workbook.save(path)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    build_clean_bom(OUT_DIR / "clean-bom.xlsx")
    build_messy_bom(OUT_DIR / "messy-bom.xlsx")
    (OUT_DIR / "rfq-email.txt").write_text(EMAIL, encoding="utf-8")

    for path in sorted(OUT_DIR.iterdir()):
        print(f"  {path.relative_to(REPO_ROOT)}  ({path.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
