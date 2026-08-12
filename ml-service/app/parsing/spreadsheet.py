"""Read a BOM out of a spreadsheet whose shape is unknown.

BOM headers are never consistent. The same column is "Part Number", "MPN",
"P/N", "Mfr Part #", "Component", or nothing at all; the table often starts
four rows down under a logo and a project name; quantities live left of the
part number as often as right.

So columns are identified by WHAT IS IN THEM, not by what the header says. Each
column is scored on the fraction of its cells that look like part numbers, like
counts, and like money. Header text is consulted only to break a tie between
two columns that scored the same on content -- which is the right precedence,
because content is what the file actually contains and the header is what
someone once typed.
"""

from __future__ import annotations

import io
import re

from .models import ParsedItem, ParseResult
from .tokens import (
    looks_like_currency,
    looks_like_quantity,
    parse_quantity,
    score_token,
    strip_edges,
)

# Scanned windows. Generous enough for a real BOM, bounded so a pathological
# sheet cannot stall the request.
MAX_ROWS = 5000
MAX_COLS = 60

# A column must be this part-number-like to be believed.
MIN_PART_COLUMN_FRACTION = 0.35

_PART_HEADERS = re.compile(
    r"\b(mpn|part\s*(number|no\.?|#)?|p/?n|manufacturer\s*part|mfr\s*part|"
    r"component|device|item\s*(number|no\.?|#)|catalog)\b",
    re.IGNORECASE,
)
_QTY_HEADERS = re.compile(
    r"\b(qty|quantity|count|pieces|pcs|order\s*qty|eau|required|amount|"
    r"units?|demand)\b",
    re.IGNORECASE,
)
_PRICE_HEADERS = re.compile(
    r"\b(price|cost|unit\s*price|target|usd|eur|gbp|each|extended)\b",
    re.IGNORECASE,
)


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # openpyxl hands back 500.0 for a cell typed as 500; rendering that as
        # "500.0" would stop it looking like a quantity.
        return str(int(value))
    return str(value).strip()


def _read_grid(sheet) -> list[list[str]]:
    """Sheet as a text grid, with merged ranges filled from their anchor.

    A merged cell reports its value only at the top-left; every other cell in
    the range comes back None. Left alone, a merged part-number cell would make
    its whole column look mostly empty and lose the column vote.
    """
    grid: list[list[str]] = []
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 1, MAX_ROWS),
                               max_col=min(sheet.max_column or 1, MAX_COLS)):
        grid.append([_cell_text(cell.value) for cell in row])

    width = max((len(row) for row in grid), default=0)
    for row in grid:
        row.extend([""] * (width - len(row)))

    # `sheet.merged_cells.ranges` directly, with no getattr fallback: Python
    # evaluates a getattr default eagerly, so a fallback to the deprecated
    # `merged_cell_ranges` ran even when it was not needed -- and that property
    # raises on this openpyxl version. The guard caused the failure it existed
    # to prevent.
    for merged in sheet.merged_cells.ranges:
        top, left = merged.min_row - 1, merged.min_col - 1
        if top >= len(grid) or left >= width:
            continue
        anchor = grid[top][left]
        if not anchor:
            continue
        for r in range(top, min(merged.max_row, len(grid))):
            for c in range(left, min(merged.max_col, width)):
                if not grid[r][c]:
                    grid[r][c] = anchor
    return grid


def _score_columns(grid: list[list[str]], skip_rows: int = 0) -> list[dict]:
    """Per-column content profile over the rows below `skip_rows`."""
    if not grid:
        return []
    width = len(grid[0])
    profiles = []

    for column in range(width):
        cells = [row[column] for row in grid[skip_rows:] if column < len(row)]
        filled = [c for c in cells if c]
        if not filled:
            profiles.append({"column": column, "filled": 0, "part": 0.0,
                             "quantity": 0.0, "currency": 0.0, "part_cells": 0})
            continue

        part_scores = [score_token(strip_edges(c)) for c in filled]
        part_cells = sum(1 for s in part_scores if s >= 0.5)
        profiles.append({
            "column": column,
            "filled": len(filled),
            "part": part_cells / len(filled),
            "part_cells": part_cells,
            "quantity": sum(1 for c in filled if looks_like_quantity(c)) / len(filled),
            "currency": sum(1 for c in filled if looks_like_currency(c)) / len(filled),
        })
    return profiles


def _find_header_row(grid: list[list[str]], part_column: int) -> int:
    """First row that holds real data in the part column.

    Everything above it is a title block, a logo row, or the header itself --
    the table simply does not start at row 1 in most real BOMs.
    """
    for index, row in enumerate(grid):
        if part_column < len(row) and score_token(strip_edges(row[part_column])) >= 0.5:
            return index
    return 0


def _header_text(grid: list[list[str]], first_data_row: int, column: int) -> str:
    """The nearest non-empty cell above the data -- the likely header label."""
    for index in range(first_data_row - 1, max(-1, first_data_row - 4), -1):
        if index < 0:
            break
        if column < len(grid[index]) and grid[index][column]:
            return grid[index][column]
    return ""


def _choose_part_column(profiles: list[dict], grid, first_guess_row: int) -> int | None:
    candidates = [p for p in profiles if p["part"] >= MIN_PART_COLUMN_FRACTION
                  and p["part_cells"] >= 2]
    if not candidates:
        # A very short BOM may only have one part row; relax the count.
        candidates = [p for p in profiles if p["part_cells"] >= 1]
    if not candidates:
        return None

    best = max(candidates, key=lambda p: (round(p["part"], 2), p["part_cells"]))
    # Header text only breaks a tie between columns that scored the same on
    # content. Content first, always.
    tied = [p for p in candidates
            if round(p["part"], 2) == round(best["part"], 2)
            and p["part_cells"] == best["part_cells"]]
    if len(tied) > 1:
        for profile in tied:
            label = _header_text(grid, first_guess_row, profile["column"])
            if _PART_HEADERS.search(label):
                return profile["column"]
    return best["column"]


def _choose_quantity_column(profiles: list[dict], grid, first_data_row: int,
                            part_column: int) -> int | None:
    candidates = [
        p for p in profiles
        if p["column"] != part_column and p["quantity"] >= 0.5 and p["filled"] >= 1
        # A column of prices is mostly decimals, but whole-dollar prices look
        # exactly like counts, so an explicit price header excludes it.
        and not _PRICE_HEADERS.search(_header_text(grid, first_data_row, p["column"]))
        and p["currency"] < 0.4
    ]
    if not candidates:
        return None

    # Prefer a column whose header says quantity; otherwise the one nearest to
    # the right of the part number, which is where BOMs usually put it.
    labelled = [p for p in candidates
                if _QTY_HEADERS.search(_header_text(grid, first_data_row, p["column"]))]
    pool = labelled or candidates
    right = [p for p in pool if p["column"] > part_column]
    return min(right or pool, key=lambda p: abs(p["column"] - part_column))["column"]


def _parse_sheet(sheet) -> tuple[ParseResult, int]:
    """Parse one worksheet. Returns the result and a score for sheet selection."""
    grid = _read_grid(sheet)
    result = ParseResult()
    if not grid:
        return result, 0

    rough = _score_columns(grid)
    part_column = _choose_part_column(rough, grid, 0)
    if part_column is None:
        result.diagnostics = {"sheet": sheet.title, "reason": "no part-number column found"}
        return result, 0

    first_data_row = _find_header_row(grid, part_column)
    # Re-score using only the data rows: a header row full of words drags a
    # column's part-number fraction down and can flip the choice.
    refined = _score_columns(grid, skip_rows=first_data_row)
    part_column = _choose_part_column(refined, grid, first_data_row) or part_column
    first_data_row = _find_header_row(grid, part_column)
    quantity_column = _choose_quantity_column(refined, grid, first_data_row, part_column)

    for index in range(first_data_row, len(grid)):
        row = grid[index]
        if part_column >= len(row):
            continue
        raw = strip_edges(row[part_column])
        if not raw:
            continue
        token_score = score_token(raw)
        if token_score < 0.5:
            if raw:
                result.skipped.append(raw)
            continue

        quantity = None
        if quantity_column is not None and quantity_column < len(row):
            quantity = parse_quantity(row[quantity_column])

        result.items.append(ParsedItem(raw, quantity, token_score, source_line=index + 1))

    result.diagnostics = {
        "parser": "spreadsheet",
        "sheet": sheet.title,
        "header_row": first_data_row,  # 0-indexed row where data starts
        "part_column": _column_letter(part_column),
        "part_column_header": _header_text(grid, first_data_row, part_column) or None,
        "quantity_column": _column_letter(quantity_column) if quantity_column is not None else None,
        "quantity_column_header": (
            _header_text(grid, first_data_row, quantity_column)
            if quantity_column is not None else None
        ) or None,
    }
    return result, len(result.items)


def _column_letter(index: int) -> str:
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def parse_spreadsheet(data: bytes, filename: str = "") -> ParseResult:
    """Parse the sheet with the most part-number-like content."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)

    best_result, best_score, considered = ParseResult(), -1, []
    for sheet in workbook.worksheets:
        try:
            result, score = _parse_sheet(sheet)
        except Exception as exc:  # noqa: BLE001 - one bad sheet is not a bad file
            considered.append({"sheet": sheet.title, "items": 0, "error": str(exc)})
            continue
        considered.append({"sheet": sheet.title, "items": score})
        if score > best_score:
            best_result, best_score = result, score

    workbook.close()

    if best_score <= 0:
        best_result.diagnostics = {
            "parser": "spreadsheet",
            "filename": filename,
            "sheets_considered": considered,
            "reason": "no sheet contained a recognisable part-number column",
        }
        return best_result

    best_result.diagnostics.update({
        "filename": filename,
        "sheets_considered": considered,
    })
    return best_result
